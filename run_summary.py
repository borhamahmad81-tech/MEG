"""
run_summary.py
---------------
Collects one row per safety event handled during a run, looks up patient
names from a user-supplied Excel/CSV list, and writes an Excel report when
the run finishes.

This module is READ-ONLY with respect to the clinical system: it never
touches the browser and never saves anything to MEG. If anything in here
fails, the run itself must continue unaffected -- every public method is
defensive for that reason.
"""

import os
from datetime import datetime

COLUMNS = [
    "Record ID",
    "Patient ID",
    "Patient Name",
    "Event Code",
    "Define the Problem",
    "Causes (Whys)",
    "Outcome",
    "Details",
]

NOT_FOUND = "Not found"
NOT_OPENED = "-"


def normalize_id(value) -> str:
    """IDs arrive from Excel as int, float or str. Normalise so that
    1000779692, '1000779692' and '1000779692.0' all match each other."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in ("nan", "none"):
        return ""
    # pandas reads a column of numeric IDs as float -> '1000779692.0'
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _normalize_header(name) -> str:
    return " ".join(str(name).strip().lower().split())


class PatientLookup:
    """ID -> Name map loaded from the user's patient list file."""

    def __init__(self, logger=None):
        self.map = {}
        self.logger = logger or (lambda *a, **k: None)
        self.source_path = ""

    @property
    def is_loaded(self) -> bool:
        return bool(self.map)

    def load(self, path: str) -> bool:
        """Load the patient list. Returns True on success. Never raises --
        a bad patient file must not stop a run, it just means names show
        as 'Not found'."""
        if not path:
            return False
        if not os.path.exists(path):
            self.logger(
                f"Patient list not found at '{path}'. Names will show as '{NOT_FOUND}'.",
                "warn",
            )
            return False

        try:
            import pandas as pd

            if path.lower().endswith(".csv"):
                df = pd.read_csv(path, dtype=str)
            else:
                df = pd.read_excel(path, sheet_name=0, dtype=str)
        except Exception as exc:
            self.logger(
                f"Could not read the patient list ({type(exc).__name__}: {exc}). "
                f"Names will show as '{NOT_FOUND}'.",
                "warn",
            )
            return False

        headers = {_normalize_header(c): c for c in df.columns}
        id_col = headers.get("patient id")
        name_col = headers.get("patient name")

        if id_col is None or name_col is None:
            self.logger(
                "The patient list must have columns named 'Patient ID' and "
                f"'Patient Name' (found: {list(df.columns)}). "
                f"Names will show as '{NOT_FOUND}'.",
                "warn",
            )
            return False

        duplicates = 0
        for _, row in df.iterrows():
            pid = normalize_id(row.get(id_col))
            name = row.get(name_col)
            name = "" if name is None else str(name).strip()
            if not pid or name.lower() in ("nan", ""):
                continue
            if pid in self.map:
                duplicates += 1
                continue  # first entry wins, as documented in the template
            self.map[pid] = name

        self.source_path = path
        message = f"Patient list loaded: {len(self.map)} patient(s) from {os.path.basename(path)}."
        if duplicates:
            message += f" ({duplicates} duplicate ID(s) ignored, first entry kept.)"
        self.logger(message, "success")
        return True

    def name_for(self, patient_id: str) -> str:
        pid = normalize_id(patient_id)
        if not pid:
            return NOT_OPENED
        if not self.map:
            return NOT_FOUND
        return self.map.get(pid, NOT_FOUND)


class RunSummary:
    """Accumulates a row per record, then writes the report."""

    def __init__(self, logger=None):
        self.logger = logger or (lambda *a, **k: None)
        self.lookup = PatientLookup(logger=self.logger)
        self.rows = []
        self.started_at = datetime.now()

    # -------------------------------------------------------------- #
    def load_patient_list(self, path: str):
        self.lookup.load(path)

    def add(self, record_id, outcome, patient_id="", event_code="",
            define_text="", causes=None, details=""):
        """Add one row. Wrapped defensively -- reporting must never be able
        to interrupt the actual run."""
        try:
            if causes is None:
                causes_text = ""
            elif isinstance(causes, (list, tuple)):
                causes_text = " | ".join(str(c) for c in causes if str(c).strip())
            else:
                causes_text = str(causes)

            pid = normalize_id(patient_id)
            self.rows.append({
                "Record ID": str(record_id or ""),
                "Patient ID": pid or NOT_OPENED,
                "Patient Name": self.lookup.name_for(pid),
                "Event Code": str(event_code or ""),
                "Define the Problem": str(define_text or ""),
                "Causes (Whys)": causes_text,
                "Outcome": str(outcome or ""),
                "Details": str(details or ""),
            })
        except Exception as exc:
            self.logger(f"Could not add a row to the run summary: {exc}", "warn")

    # -------------------------------------------------------------- #
    def counts_by_outcome(self):
        counts = {}
        for row in self.rows:
            key = row["Outcome"]
            counts[key] = counts.get(key, 0) + 1
        return counts

    def export(self, folder: str = "", heading: str = "") -> str:
        """Write the summary to an .xlsx file. Returns the path written, or
        '' if nothing was written. Never raises."""
        if not self.rows:
            self.logger("No records were handled, so no summary file was written.", "info")
            return ""

        try:
            folder = folder or os.getcwd()
            os.makedirs(folder, exist_ok=True)
            stamp = self.started_at.strftime("%Y-%m-%d_%H-%M")
            path = os.path.join(folder, f"run_summary_{stamp}.xlsx")
            path = self._unique_path(path)
            self._write_workbook(path, heading)
            self.logger(f"Summary report saved: {path}", "success")
            return path
        except Exception as exc:
            self.logger(
                f"Could not write the summary report ({type(exc).__name__}: {exc}). "
                "The run itself was not affected.",
                "warn",
            )
            return ""

    @staticmethod
    def _unique_path(path: str) -> str:
        if not os.path.exists(path):
            return path
        stem, ext = os.path.splitext(path)
        n = 2
        while os.path.exists(f"{stem}_{n}{ext}"):
            n += 1
        return f"{stem}_{n}{ext}"

    def _write_workbook(self, path: str, heading: str):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Run Summary"

        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E8B57")
        body_font = Font(name="Arial", size=10)

        for col, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r, row in enumerate(self.rows, start=2):
            for c, name in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=r, column=c, value=row.get(name, ""))
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=(name in (
                    "Define the Problem", "Causes (Whys)", "Details")))

        widths = {
            "Record ID": 12, "Patient ID": 15, "Patient Name": 26,
            "Event Code": 12, "Define the Problem": 48,
            "Causes (Whys)": 48, "Outcome": 20, "Details": 34,
        }
        for col, name in enumerate(COLUMNS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = widths.get(name, 18)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(COLUMNS)).column_letter}{len(self.rows) + 1}"

        # ---- second sheet: run info ----
        info = wb.create_sheet("Run Info")
        counts = self.counts_by_outcome()
        lines = [
            ("Safety Event Assistant - Run Summary", True),
            ("", False),
            (f"Run started : {self.started_at.strftime('%Y-%m-%d %H:%M')}", False),
            (f"Report written : {datetime.now().strftime('%Y-%m-%d %H:%M')}", False),
            (f"Result : {heading}" if heading else "", False),
            (f"Total records listed : {len(self.rows)}", False),
            ("", False),
            ("Breakdown by outcome", True),
        ]
        for outcome, count in sorted(counts.items()):
            lines.append((f"{outcome} : {count}", False))
        lines += [
            ("", False),
            ("Patient list", True),
            (f"Source : {self.lookup.source_path or 'not loaded'}", False),
            (f"Patients in list : {len(self.lookup.map)}", False),
            ("", False),
            ("Notes", True),
            ("'Serious' events are never opened by this app - they are listed here", False),
            ("for your awareness only and must be handled manually.", False),
            (f"'{NOT_OPENED}' in a patient column means the record was not opened,", False),
            ("so no patient ID could be read from it.", False),
            (f"'{NOT_FOUND}' means the record's patient ID was not in your patient list.", False),
        ]
        r = 1
        for text, bold in lines:
            cell = info.cell(row=r, column=1, value=text)
            cell.font = Font(name="Arial", size=11, bold=bold,
                             color="2E8B57" if bold else "000000")
            r += 1
        info.column_dimensions["A"].width = 82

        wb.save(path)
